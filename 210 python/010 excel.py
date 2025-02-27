pip freeze --all  # see installed packages
pip list  # freeze with a nicer format
pip install openpyxl  # install in cmd
import openpyxl  # if no errors in python console = install ok
import os

''' ---------------------------- DEFINITIONS  --------------------------  '''
workbook  # excel file
sheets/worksheets  # excel tab
Columns/Rows/Cells 

''' ---------------------------- OPEN FILES ----------------------------  '''
# get path to file
rel_path = os.path.join('..', 'data', 'example.xlsx')
abs_path = os.path.abspath(rel_path)

# get data from excel
workbook = openpyxl.load_workbook(abs_path)
workbook.get_sheet_names() 
sheet = workbook.get_sheet_by_name('Sheet1')
cell = sheet['A1']  # returns cell A1 as object
cell = sheet.cell(row=1, column=1)  # the same
A1_value = cell.value


# quick way
A1_value = str(sheet['A1'].value)  # for cell with datetime
A2_value = sheet['B1'].value

# get data from cell in a column
for i in range(1,8):
  print(i, sheet.cell(row=i, column=2).value)

''' --------------------------- CREATE FILES ---------------------------  '''
wb = openpyxl.Workbook()  # creates Workbook object with a single sheet named Sheet
sheet = wb.get_sheet_by_name('Sheet')
sheet['A1'] = 'new data in sheet'
wb.save('example2.xlsx')

# create sheet an save it in object for later use
sheet2 = wb.create_sheet() # it takes the default name 'Sheet1'
sheet2.title = 'New Name For Sheet'
wb.remove_sheet(wb.get_sheet_by_name('Sheet'))  # remove default sheet
#quickway
wb.create_sheet(index=0, title='MyName')  # first position to the left

''' -------------------------- FIND EMPTY CELLS ------------------------  '''
# exit if row has no info (value doesn't exist or value == None)
  if not hasattr(sheet1['B'+str(row)], 'value')\
          or not sheet1['B'+str(row)].value:
      break


